"""サンプル結合テーブル(combined)を**読みやすく整形した xlsx** として書き出す。

`combined.to_excel()` の素の出力（列幅も色も無し・日付に 00:00:00 が出る）を置き換える。

構成:
  ・Sheet「sample_table」… データ本体。列幅/行高/ヘッダ色(ジャンル別)/枠固定/フィルタ/日付書式
  ・Sheet「カラム説明」  … 列ごとの ジャンル/型/非欠損/充足率/略/内容（columns.COLUMN_DOCS + 実測）

使い方:
    import xlsx_report
    xlsx_report.build_xlsx(combined, "sample_table.xlsx")

方針: **データセルには個別スタイルを付けない**。9364×73 ≈ 68万セルを個別に塗ると
ファイルが肥大化して書き出しも Excel での描画も遅くなるため、色はヘッダ行だけに乗せ、
書式は「列単位」でまとめて当てる。
"""
from __future__ import annotations

import unicodedata

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import columns as coldocs
from table_report import _genre

SHEET_DATA = "sample_table"
SHEET_DOCS = "カラム説明"

# 列幅（表示文字数。全角は2カウント）
WIDTH_MIN, WIDTH_MAX = 6, 40
WIDTH_SAMPLE_ROWS = 400        # 幅算出に見る先頭行数（全行なめると遅い）
WIDTH_PAD = 2

ROW_HEIGHT = 15                # データ行の高さ（全行統一）
HEADER_HEIGHT = 30             # ヘッダ行だけ少し高く（2行折り返し用）

# 列ごとの表示書式（列名 → Excel number_format）
DATE_FMT = "YYYY-MM-DD"
_FMT_BY_COL = {
    "birth": DATE_FMT,
    "death": DATE_FMT,
    "age": "0",
    "DNA_conc": "0.000",
}
_FMT_BY_PREFIX = {"bv_": "0.000"}


def _disp_len(v) -> int:
    """全角を2、半角を1として表示幅を数える。"""
    s = "" if v is None else str(v)
    return sum(2 if unicodedata.east_asian_width(ch) in "WFA" else 1 for ch in s)


def _fmt_for(col: str) -> str | None:
    if col in _FMT_BY_COL:
        return _FMT_BY_COL[col]
    for p, f in _FMT_BY_PREFIX.items():
        if col.startswith(p):
            return f
    return None


def _col_width(df: pd.DataFrame, col: str) -> float:
    """ヘッダ名と先頭 WIDTH_SAMPLE_ROWS 行の中身から表示幅を決める。"""
    head = df[col].head(WIDTH_SAMPLE_ROWS)
    if _fmt_for(col) == DATE_FMT:
        body = 10                                   # YYYY-MM-DD 固定
    else:
        body = max((_disp_len(v) for v in head.dropna()), default=0)
    w = max(_disp_len(col), body) + WIDTH_PAD
    return float(min(max(w, WIDTH_MIN), WIDTH_MAX))


def _write_data_sheet(ws, df: pd.DataFrame) -> None:
    """データ本体シート（値はすでに書き込み済み）にスタイルを当てる。"""
    thin = Side(style="thin", color="FFFFFF")

    # --- ヘッダ行: ジャンル別に塗り分け + 太字白文字 + 折り返し中央 ---
    for j, col in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=j)
        fill = coldocs.GENRE_COLORS[_genre(col)]["header"]
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=thin, right=thin)

    # --- 列幅 ---
    for j, col in enumerate(df.columns, 1):
        ws.column_dimensions[get_column_letter(j)].width = _col_width(df, col)

    # --- 行高: 既定値で全行統一。ヘッダだけ個別に高く ---
    ws.sheet_format.defaultRowHeight = ROW_HEIGHT
    ws.sheet_format.customHeight = True
    ws.row_dimensions[1].height = HEADER_HEIGHT

    # --- 表示書式（列単位でまとめて当てる） ---
    n = len(df)
    for j, col in enumerate(df.columns, 1):
        fmt = _fmt_for(col)
        if not fmt:
            continue
        letter = get_column_letter(j)
        for i in range(2, n + 2):
            ws[f"{letter}{i}"].number_format = fmt

    # --- 枠固定 / オートフィルタ ---
    # name(3列目) までを固定して、横スクロールしても個体が分かるようにする
    ws.freeze_panes = "D2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{n + 1}"


def _write_docs_sheet(ws, df: pd.DataFrame) -> None:
    """カラム説明シートを組む。"""
    headers = ["#", "列名", "ジャンル", "型", "非欠損", "充足率", "略", "内容"]
    widths = [5, 20, 12, 10, 9, 8, 34, 62]

    ws.append(headers)
    for j, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=j)
        c.fill = PatternFill("solid", fgColor="404040")
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(j)].width = w

    for r in coldocs.describe_columns(df, plain_text=True):     # セルにバックティックは出さない
        ws.append([r["no"], r["col"], r["genre"], r["dtype"],
                   r["n_notna"], r["fill_rate"], r["abbr"], r["desc"]])
        i = ws.max_row
        light = coldocs.GENRE_COLORS[r["genre"]]["light"]
        for j in range(1, len(headers) + 1):                # ジャンル色は淡色で行全体に
            ws.cell(row=i, column=j).fill = PatternFill("solid", fgColor=light)
        ws.cell(row=i, column=2).font = Font(bold=True, size=10)
        ws.cell(row=i, column=6).number_format = "0.0%"
        ws.cell(row=i, column=5).number_format = "#,##0"
        ws.cell(row=i, column=8).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=i, column=7).alignment = Alignment(wrap_text=True, vertical="top")

    ws.sheet_format.defaultRowHeight = ROW_HEIGHT
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"


def build_xlsx(combined: pd.DataFrame, path) -> None:
    """combined を整形済み xlsx として path に書き出す（データ + カラム説明の2シート）。"""
    with pd.ExcelWriter(path, engine="openpyxl") as xw:
        combined.to_excel(xw, sheet_name=SHEET_DATA, index=False)
        _write_data_sheet(xw.sheets[SHEET_DATA], combined)

        ws_docs = xw.book.create_sheet(SHEET_DOCS)
        _write_docs_sheet(ws_docs, combined)
